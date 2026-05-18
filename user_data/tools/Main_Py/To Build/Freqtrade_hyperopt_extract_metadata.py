from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import tkinter as tk
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, ttk
from typing import Any, Iterable, List, Optional, Tuple

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore

    HAS_DND = True
except Exception:
    DND_FILES = None
    TkinterDnD = None
    HAS_DND = False


########################################################################################################################################################
# Output columns - keep this order EXACTLY the same as your Excel Hyperopt_Log columns.
########################################################################################################################################################

COLUMNS: List[str] = [
    "Strategy",
    "Created At",
    "Hyperopt Seed",
    "Epochs Count",
    "Strategy Version",
    "Hyperopt Loss",
    "Timerange Group",
    "Train Timerange",
    "Valid Timerange",
    "Test Timerange",
    "Lookahead Checked",
    "Lookahead Timerange",
    "Lookahead Has Bias",
    "Lookahead Total Signals",
    "Recursive Checked",
    "Recursive Timerange",
    "Recursive Run Status",
    "Recursive Pair Count",
    "Recursive Finished Pairs",
    "Recursive OK Pairs",
    "Recursive Failed Pairs",
    "Recursive Pending Pairs",
    "Recursive Max Abs Variance %",
]

ANSI_RE = re.compile(r"\x1b\[[0-?]*[ -/]*[@-~]")

SKIP_FAILED_REASONS = {
    "NO_DATA",
    "NO_DATA_AVAILABLE",
    "DATA_NOT_AVAILABLE",
    "MISSING_DATA",
    "EMPTY_DATA",
    "EMPTY_DATAFRAME",
    "NO_PAIR_IN_WHITELIST",
    "PAIR_NOT_IN_WHITELIST",
    "NOT_IN_WHITELIST",
    "PAIR_NOT_AVAILABLE",
    "PAIR_UNAVAILABLE",
    "NO_PAIR",
    "NO_TRADES",
    "STALE_JOB",
    "SKIPPED",
    "SKIP",
}


########################################################################################################################################################
# Parser core
########################################################################################################################################################

def clean_text(text: str) -> str:
    text = ANSI_RE.sub("", text or "")
    text = text.replace("\ufeff", "")
    text = text.replace("−", "-").replace("\xa0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text


def read_text_file(filepath: str | Path) -> str:
    path = Path(str(filepath).strip())

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    for enc in ["utf-8", "utf-8-sig", "cp1252", "latin-1"]:
        try:
            return path.read_text(encoding=enc)
        except Exception:
            pass

    raise ValueError(f"Could not read file as text: {path}")


def kv(text: str, key: str) -> Optional[str]:
    key_pattern = re.escape(key).replace(r"\*", r"[*\s-]")
    match = re.search(rf"(?im)^\s*{key_pattern}\s*[:=]\s*(.*?)\s*$", text)
    return match.group(1).strip() if match else None


def first_non_empty(*values: Any) -> str:
    for value in values:
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return ""


def parse_bool(raw: Optional[str]) -> Optional[bool]:
    if raw is None:
        return None

    value = raw.strip().lower()

    if value in {"true", "yes", "y", "1", "pass", "passed", "ok", "clean"}:
        return True

    if value in {"false", "no", "n", "0", "fail", "failed", "missing"}:
        return False

    return None


def bool_for_excel(value: Optional[bool]) -> str:
    if value is True:
        return "TRUE"
    if value is False:
        return "FALSE"
    return ""


def normalize_yes_no(value: Any) -> str:
    raw = str(value or "").strip()

    if not raw:
        return ""

    low = raw.lower()

    if low in {"no", "false", "0", "clean", "none"}:
        return "No"

    if low in {"yes", "true", "1", "bias", "biased"}:
        return "Yes"

    return raw


def safe_int(value: Any) -> int | str:
    raw = str(value or "").strip()

    if not raw:
        return ""

    match = re.search(r"-?\d+", raw)

    if not match:
        return raw

    try:
        return int(match.group(0))
    except ValueError:
        return raw


def safe_float(value: Any) -> float | str:
    raw = str(value or "").strip().replace("%", "")

    if not raw:
        return ""

    match = re.search(r"-?\d+(?:\.\d+)?", raw)

    if not match:
        return raw

    try:
        return float(match.group(0))
    except ValueError:
        return raw


def normalize_strategy_key(value: Any) -> str:
    raw = str(value or "").strip()
    raw = Path(raw).stem if raw.lower().endswith(".py") else raw
    raw = raw.lower()
    raw = re.sub(r"[^a-z0-9]+", "", raw)
    return raw


def normalize_status_token(value: Any) -> str:
    raw = str(value or "").strip().upper()
    raw = re.sub(r"[^A-Z0-9]+", "_", raw)
    return raw.strip("_")


def strategy_version_from_name(strategy: str, text: str) -> str:
    explicit = first_non_empty(
        kv(text, "strategy_version"),
        kv(text, "strategy version"),
        kv(text, "version"),
    )

    if explicit:
        return explicit

    match = re.search(r"(?:[_\-]v?|V)(\d+(?:[_\.]\d+)*)$", strategy or "")
    return match.group(1) if match else ""


def created_at_value(text: str) -> str:
    raw = first_non_empty(
        kv(text, "created_at"),
        kv(text, "created at"),
        kv(text, "updated_at"),
        kv(text, "updated at"),
    )

    if not raw:
        raw_file = first_non_empty(
            kv(text, "raw_output_file"),
            kv(text, "report_file"),
            kv(text, "combined_raw_output_file"),
        )

        if raw_file:
            match = re.search(r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})", raw_file)
            raw = match.group(1) if match else ""

    if not raw:
        return ""

    for fmt in ("%Y-%m-%d_%H-%M-%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            pass

    return raw


def normalize_timerange(value: Any) -> str:
    raw = str(value or "").strip()

    if not raw:
        return ""

    compact = re.search(r"(\d{8})\s*-\s*(\d{8})", raw)
    if compact:
        return f"{compact.group(1)}-{compact.group(2)}"

    dated = re.search(
        r"(\d{4})-(\d{2})-(\d{2})(?:\s+\d{2}:\d{2}:\d{2})?\s*(?:->|to|-)\s*"
        r"(\d{4})-(\d{2})-(\d{2})(?:\s+\d{2}:\d{2}:\d{2})?",
        raw,
        flags=re.IGNORECASE,
    )
    if dated:
        return f"{dated.group(1)}{dated.group(2)}{dated.group(3)}-{dated.group(4)}{dated.group(5)}{dated.group(6)}"

    return raw


def window_timerange(text: str, window_name: str) -> str:
    match = re.search(
        rf"(?im)^=+\s*WINDOW\s+{re.escape(window_name)}\s*\|\s*TIMERANGE\s+([0-9]{{8}}-[0-9]{{8}})\s*=+",
        text,
    )
    return match.group(1).strip() if match else ""


def all_window_timeranges(text: str) -> List[str]:
    ranges: List[str] = []

    for match in re.finditer(
        r"(?im)^=+\s*WINDOW\s+[^|=]+\|\s*TIMERANGE\s+([0-9]{8}-[0-9]{8})\s*=+",
        text,
    ):
        ranges.append(match.group(1).strip())

    return ranges


def backtested_timeranges(text: str) -> List[str]:
    ranges: List[str] = []

    for match in re.finditer(
        r"(?im)^\s*Backtested\s+(\d{4})-(\d{2})-(\d{2})\s+\d{2}:\d{2}:\d{2}\s*->\s*(\d{4})-(\d{2})-(\d{2})\s+\d{2}:\d{2}:\d{2}",
        text,
    ):
        ranges.append(
            f"{match.group(1)}{match.group(2)}{match.group(3)}-"
            f"{match.group(4)}{match.group(5)}{match.group(6)}"
        )

    return ranges


def combined_minmax_timerange(ranges: Iterable[str]) -> str:
    starts: List[str] = []
    ends: List[str] = []

    for value in ranges:
        normalized = normalize_timerange(value)
        match = re.match(r"^(\d{8})-(\d{8})$", normalized)
        if match:
            starts.append(match.group(1))
            ends.append(match.group(2))

    if not starts or not ends:
        return ""

    return f"{min(starts)}-{max(ends)}"


def timerange_from_text(text: str, key_prefix: str = "") -> str:
    prefix_keys: List[str] = []

    if key_prefix:
        prefix_keys = [
            f"{key_prefix}_timerange",
            f"{key_prefix} timerange",
            f"{key_prefix} time range",
            f"{key_prefix}_time_range",
        ]

    candidates = [kv(text, key) for key in prefix_keys]
    candidates.extend(
        [
            kv(text, "timerange"),
            kv(text, "analysis_timerange"),
            kv(text, "analysis timerange"),
            kv(text, "time range"),
        ]
    )

    for candidate in candidates:
        normalized = normalize_timerange(candidate)
        if normalized:
            return normalized

    window_combined = combined_minmax_timerange(all_window_timeranges(text))
    if window_combined:
        return window_combined

    backtested_combined = combined_minmax_timerange(backtested_timeranges(text))
    if backtested_combined:
        return backtested_combined

    return ""


def epochs_count(text: str) -> int | str:
    explicit = first_non_empty(
        kv(text, "epochs_count"),
        kv(text, "epochs count"),
        kv(text, "trial_count"),
        kv(text, "trial count"),
        kv(text, "epochs"),
    )

    if explicit:
        match = re.search(r"\d+", explicit)
        if match:
            return int(match.group(0))
        return explicit

    candidates: List[int] = []

    patterns = [
        r"(?im)^Epochs\s+.*?(\d+)\s*/\s*(\d+)",
        r"(?im)^\s*(\d+)\s*/\s*(\d+)\s*:",
        r"│\s*\*?\s*Best\s*│\s*(\d+)\s*/\s*(\d+)\s*│",
        r"┃\s*Best\s*┃\s*Epoch\s*┃.*?\n.*?│\s*\*?\s*Best\s*│\s*(\d+)\s*/\s*(\d+)",
        r"(?im)(\d+)\s+epochs\s+saved",
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.DOTALL):
            try:
                if len(match.groups()) >= 2:
                    candidates.append(int(match.group(2)))
                else:
                    candidates.append(int(match.group(1)))
            except (ValueError, IndexError):
                pass

    return max(candidates) if candidates else ""


def empty_output_row() -> dict[str, Any]:
    return {col: "" for col in COLUMNS}


def unique_paths(paths: Iterable[str]) -> List[str]:
    output: List[str] = []
    seen: set[str] = set()

    for path in paths:
        clean_path = str(Path(path))
        key = str(Path(clean_path).resolve()) if Path(clean_path).exists() else clean_path.lower()

        if key in seen:
            continue

        seen.add(key)
        output.append(clean_path)

    return output


########################################################################################################################################################
# File type detection
########################################################################################################################################################

def is_hyperopt_text(text: str) -> bool:
    text = clean_text(text)

    return bool(
        re.search(r"(?i)#\s*Freqtrade\s+Hyperopt\s+Extract\s+Metadata", text)
        or kv(text, "hyperopt_loss")
        or kv(text, "hyperopt loss")
        or kv(text, "random_state")
        or kv(text, "hyperopt_seed")
        or re.search(r"(?i)Hyperopt results", text)
        or re.search(r"(?i)Best result", text)
        or re.search(r"(?i)Using Hyperopt loss class name", text)
        or re.search(r"(?i)Using resolved hyperoptloss", text)
    )


def is_lookahead_text(text: str) -> bool:
    text = clean_text(text)

    return bool(
        re.search(r"(?im)^\s*mode\s*=\s*lookahead-analysis\s*$", text)
        or re.search(r"(?im)^\s*analysis\s*=\s*Lookahead-Analysis\s*$", text)
        or re.search(r"(?im)^\s*Lookahead Analysis\s*$", text)
        or re.search(r"(?i)#\s*Freqtrade\s+Lookahead", text)
        or re.search(r"(?i)lookahead-analysis", text)
        or kv(text, "lookahead_timerange")
        or kv(text, "lookahead total signals")
        or kv(text, "lookahead_total_signals")
        or kv(text, "lookahead_has_bias")
    )


def is_recursive_text(text: str) -> bool:
    text = clean_text(text)

    return bool(
        re.search(r"(?im)^\s*mode\s*=\s*recursive-analysis\s*$", text)
        or re.search(r"(?im)^\s*analysis\s*=\s*Recursive-Analysis\s*$", text)
        or re.search(r"(?i)#\s*Freqtrade Recursive-Analysis Report", text)
        or re.search(r"(?i)Recursive-Analysis Tables", text)
        or re.search(r"(?im)^\s*Recursive Analysis\s*$", text)
        or re.search(r"(?i)recursive-analysis", text)
        or kv(text, "recursive_timerange")
        or kv(text, "recursive max abs variance")
        or kv(text, "recursive_max_abs_variance")
    )


def detected_kinds_for_text(text: str) -> List[str]:
    clean = clean_text(text)

    # Hyperopt wins as a standalone kind because hyperopt extracts can contain lookahead/recursive words in commands/notes.
    if is_hyperopt_text(clean):
        return ["hyperopt"]

    kinds: List[str] = []

    if is_lookahead_text(clean):
        kinds.append("lookahead")

    if is_recursive_text(clean):
        kinds.append("recursive")

    return kinds or ["hyperopt"]


########################################################################################################################################################
# Hyperopt parser
########################################################################################################################################################

def parse_hyperopt_extract_text(raw_text: str) -> dict[str, Any]:
    text = clean_text(raw_text)

    strategy = first_non_empty(kv(text, "strategy"), kv(text, "Strategy"))

    if not strategy:
        match = re.search(r"/strategy_([A-Za-z0-9_]+)_\d{4}-\d{2}-\d{2}", text)
        if match:
            strategy = match.group(1)

    if not strategy:
        match = re.search(r"│\s*([A-Za-z_][A-Za-z0-9_]*)\s*│\s*\d+\s*│", text)
        if match:
            strategy = match.group(1)

    timerange_group = first_non_empty(
        kv(text, "time_window"),
        kv(text, "timerange_group"),
        kv(text, "timerange group"),
    ).upper()

    metadata_timerange = normalize_timerange(
        first_non_empty(
            kv(text, "timerange"),
            kv(text, "train_timerange"),
            kv(text, "train timerange"),
        )
    )

    train_range = window_timerange(text, "TRAIN")
    valid_range = window_timerange(text, "VALID")
    test_range = window_timerange(text, "TEST")

    if timerange_group == "TRAIN" and not train_range:
        train_range = metadata_timerange
    elif timerange_group == "VALID" and not valid_range:
        valid_range = metadata_timerange
    elif timerange_group == "TEST" and not test_range:
        test_range = metadata_timerange
    elif timerange_group in {"FULL", "CUSTOM", "LIVE_CHECK"} and not train_range:
        train_range = metadata_timerange

    fallback_ranges = backtested_timeranges(text)
    used_ranges = {x for x in [train_range, valid_range, test_range] if x}
    remaining_ranges = [x for x in fallback_ranges if x not in used_ranges]

    if not train_range and metadata_timerange and timerange_group in {"TRAIN", "FULL", "CUSTOM", "LIVE_CHECK", ""}:
        train_range = metadata_timerange

    if not train_range and len(remaining_ranges) >= 3:
        train_range = remaining_ranges.pop(0)

    if not valid_range and remaining_ranges:
        valid_range = remaining_ranges.pop(0)

    if not test_range and remaining_ranges:
        test_range = remaining_ranges.pop(0)

    seed_raw = first_non_empty(
        kv(text, "random_state"),
        kv(text, "hyperopt_seed"),
        kv(text, "hyperopt seed"),
    )

    seed: int | str = ""

    if seed_raw:
        try:
            seed = int(seed_raw)
        except ValueError:
            seed = seed_raw

    hyperopt_loss = first_non_empty(kv(text, "hyperopt_loss"), kv(text, "hyperopt loss"))

    if not hyperopt_loss:
        match = re.search(r"(?im)Using Hyperopt loss class name:\s*([A-Za-z_][A-Za-z0-9_]*)", text)
        if match:
            hyperopt_loss = match.group(1)

    if not hyperopt_loss:
        match = re.search(r"(?im)Using resolved hyperoptloss\s+([A-Za-z_][A-Za-z0-9_]*)", text)
        if match:
            hyperopt_loss = match.group(1)

    row = empty_output_row()
    row.update(
        {
            "Strategy": strategy,
            "Created At": created_at_value(text),
            "Hyperopt Seed": seed,
            "Epochs Count": epochs_count(text),
            "Strategy Version": strategy_version_from_name(strategy, text),
            "Hyperopt Loss": hyperopt_loss,
            "Timerange Group": timerange_group,
            "Train Timerange": train_range,
            "Valid Timerange": valid_range,
            "Test Timerange": test_range,
            "Lookahead Timerange": "",
            "Recursive Timerange": "",
            "Lookahead Checked": "",
            "Lookahead Has Bias": "",
            "Lookahead Total Signals": "",
            "Recursive Checked": "",
            "Recursive Run Status": "",
            "Recursive Pair Count": "",
            "Recursive Finished Pairs": "",
            "Recursive OK Pairs": "",
            "Recursive Failed Pairs": "",
            "Recursive Pending Pairs": "",
            "Recursive Max Abs Variance %": "",
        }
    )

    return {col: row.get(col, "") for col in COLUMNS}


def parse_hyperopt_extract_file(path: str | Path) -> dict[str, Any]:
    return parse_hyperopt_extract_text(read_text_file(path))


########################################################################################################################################################
# Lookahead parser
########################################################################################################################################################

def parse_lookahead_text(raw_text: str) -> dict[str, Any]:
    text = clean_text(raw_text)

    strategy = first_non_empty(kv(text, "strategy"), kv(text, "Strategy"))
    lookahead_timerange = timerange_from_text(text, "lookahead")

    has_bias = first_non_empty(
        kv(text, "has_bias"),
        kv(text, "lookahead_has_bias"),
        kv(text, "lookahead has bias"),
    )

    total_signals = first_non_empty(
        kv(text, "total_signals"),
        kv(text, "lookahead_total_signals"),
        kv(text, "lookahead total signals"),
        kv(text, "total signals"),
    )

    table_match = re.search(
        r"[│┃]\s*([^│┃]+?\.py)\s*[│┃]\s*([A-Za-z_][A-Za-z0-9_]*)\s*[│┃]\s*(Yes|No|True|False)\s*[│┃]\s*([0-9]+)\s*[│┃]",
        text,
        flags=re.IGNORECASE,
    )

    if table_match:
        if not strategy:
            strategy = table_match.group(2).strip()
        if not has_bias:
            has_bias = table_match.group(3).strip()
        if not total_signals:
            total_signals = table_match.group(4).strip()

    if not strategy:
        filename_match = re.search(r"([A-Za-z_][A-Za-z0-9_]*)\.py", text)
        if filename_match:
            strategy = filename_match.group(1).strip()

    has_bias = normalize_yes_no(has_bias)

    return {
        "strategy": strategy,
        "lookahead_timerange": lookahead_timerange,
        "lookahead_checked": "TRUE",
        "lookahead_has_bias": has_bias,
        "lookahead_total_signals": safe_int(total_signals),
        "lookahead_created_at": created_at_value(text),
    }


########################################################################################################################################################
# Recursive parser
########################################################################################################################################################

def recursive_job_summary_value(text: str, key: str) -> str:
    direct = kv(text, key)

    if direct:
        return direct

    match = re.search(rf"(?im)^\s*{re.escape(key)}\s*=\s*(\d+)\s*$", text)

    if match:
        return match.group(1).strip()

    return ""


def recursive_ok_blocks(text: str) -> List[str]:
    blocks: List[str] = []

    for match in re.finditer(
        r"(?ims)^=+\s*RUN\s+.+?\s*\|\s*OK\b.*?=+\s*(.*?)^=+\s*END RUN\s+.+?\s*=+\s*$",
        text,
    ):
        blocks.append(match.group(1))

    return blocks


def max_abs_percent_from_text(text: str) -> float | str:
    values: List[float] = []

    for match in re.finditer(r"(-?\d+(?:\.\d+)?)\s*%", text):
        try:
            values.append(abs(float(match.group(1))))
        except ValueError:
            pass

    return max(values) if values else ""


def max_abs_recursive_variance_from_ok_blocks(text: str) -> float | str:
    blocks = recursive_ok_blocks(text)

    if not blocks:
        return ""

    values: List[float] = []

    for block in blocks:
        block_value = max_abs_percent_from_text(block)

        if block_value != "":
            try:
                values.append(float(block_value))
            except ValueError:
                pass

    return max(values) if values else ""


def recursive_failed_reasons(text: str) -> List[str]:
    reasons: List[str] = []

    for line in clean_text(text).splitlines():
        # Markdown/ASCII job-summary row:
        # | Pair | FAILED | effective_code | docker_code | NO_DATA |
        if "|" not in line or not re.search(r"\bFAILED\b|\bFAIL\b|\bERROR\b", line, flags=re.IGNORECASE):
            continue

        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]

        if len(cells) < 2:
            continue

        status_cells = [normalize_status_token(cell) for cell in cells]

        if not any(cell in {"FAILED", "FAIL", "ERROR"} for cell in status_cells):
            continue

        reason = normalize_status_token(cells[-1])

        if reason and reason not in {"STATUS", "REASON", "FAILED", "FAIL", "ERROR", "NA", "N_A", "NONE", ""}:
            reasons.append(reason)

    # Fallback for logs that print reason=NO_DATA / reason: NO_DATA.
    for match in re.finditer(r"(?im)\breason\s*[:=]\s*([A-Za-z0-9_\- ]+)\s*$", text):
        reason = normalize_status_token(match.group(1))
        if reason:
            reasons.append(reason)

    return reasons


def failed_reasons_are_only_skips(failed_reasons: List[str]) -> bool:
    if not failed_reasons:
        return False

    return all(reason in SKIP_FAILED_REASONS for reason in failed_reasons)


def infer_recursive_run_status(
    run_status: str,
    pair_count: int | str,
    finished_count: int | str,
    ok_count: int | str,
    failed_count: int | str,
    skipped_count: int | str,
    pending_count: int | str,
    failed_reasons: Optional[List[str]] = None,
) -> str:
    failed_reasons = failed_reasons or []
    raw_status = normalize_status_token(run_status)

    def to_int(value: int | str) -> int:
        try:
            return int(value) if value != "" else 0
        except Exception:
            return 0

    pair_i = to_int(pair_count)
    finished_i = to_int(finished_count)
    ok_i = to_int(ok_count)
    failed_i = to_int(failed_count)
    skipped_i = to_int(skipped_count)
    pending_i = to_int(pending_count)

    if raw_status in {"IN_PROGRESS", "RUNNING", "STARTED"}:
        return "IN_PROGRESS"

    if pending_i > 0:
        return "IN_PROGRESS"

    if pair_i > 0 and finished_i < pair_i:
        return "IN_PROGRESS"

    # Important: wrapper reports can call NO_DATA whitelist/data skips "FAILED".
    # For your sheet these should not become technical warnings.
    if failed_i > 0:
        if failed_reasons_are_only_skips(failed_reasons):
            return "FINISHED_WITH_SKIPPED"
        return "FINISHED_WITH_TECHNICAL_WARNINGS"

    if skipped_i > 0:
        return "FINISHED_WITH_SKIPPED"

    if ok_i > 0:
        return "FINISHED_OK"

    if raw_status in {"FINISHED_OK", "FINISHED", "SUCCESS", "PASSED", "PASS"}:
        return "FINISHED_OK"

    if raw_status in {"FINISHED_WITH_SKIPPED", "FINISHED_WITH_SKIPS", "PASS_WITH_SKIPS"}:
        return "FINISHED_WITH_SKIPPED"

    if raw_status in {"FINISHED_WITH_FAILURES", "FINISHED_WITH_FAILS", "FAILED_WITH_NO_DATA"}:
        if failed_reasons_are_only_skips(failed_reasons):
            return "FINISHED_WITH_SKIPPED"
        return "FINISHED_WITH_TECHNICAL_WARNINGS"

    if pair_i > 0 and finished_i >= pair_i and ok_i <= 0:
        return "FAILED"

    return raw_status if raw_status else ""


def parse_recursive_text(raw_text: str) -> dict[str, Any]:
    text = clean_text(raw_text)

    strategy = first_non_empty(kv(text, "strategy"), kv(text, "Strategy"))
    recursive_timerange = timerange_from_text(text, "recursive")

    run_status = first_non_empty(kv(text, "run_status"), kv(text, "run status"))

    pair_count = safe_int(
        first_non_empty(
            kv(text, "pair_count"),
            kv(text, "pair count"),
            kv(text, "total_pairs"),
            kv(text, "total pairs"),
            recursive_job_summary_value(text, "total_pairs"),
        )
    )

    finished_count = safe_int(
        first_non_empty(
            kv(text, "finished_count"),
            kv(text, "finished count"),
            recursive_job_summary_value(text, "finished"),
        )
    )

    ok_count = safe_int(
        first_non_empty(
            kv(text, "ok_count"),
            kv(text, "ok count"),
            recursive_job_summary_value(text, "ok"),
        )
    )

    skipped_count = safe_int(
        first_non_empty(
            kv(text, "skipped_count"),
            kv(text, "skipped count"),
            recursive_job_summary_value(text, "skipped"),
        )
    )

    failed_count = safe_int(
        first_non_empty(
            kv(text, "technical_failed_count"),
            kv(text, "technical failed count"),
            kv(text, "failed_count"),
            kv(text, "failed count"),
            recursive_job_summary_value(text, "technical_failed"),
            recursive_job_summary_value(text, "failed"),
        )
    )

    pending_count = safe_int(
        first_non_empty(
            kv(text, "pending_count"),
            kv(text, "pending count"),
            recursive_job_summary_value(text, "pending"),
        )
    )

    failed_reasons = recursive_failed_reasons(text)

    run_status = infer_recursive_run_status(
        run_status=run_status,
        pair_count=pair_count,
        finished_count=finished_count,
        ok_count=ok_count,
        failed_count=failed_count,
        skipped_count=skipped_count,
        pending_count=pending_count,
        failed_reasons=failed_reasons,
    )

    explicit_max_abs_variance = first_non_empty(
        kv(text, "recursive_max_abs_variance"),
        kv(text, "recursive max abs variance"),
        kv(text, "recursive max abs variance %"),
        kv(text, "max_abs_variance"),
        kv(text, "max abs variance"),
        kv(text, "max abs variance %"),
    )

    if explicit_max_abs_variance:
        max_abs_variance_value = safe_float(explicit_max_abs_variance)
    else:
        max_abs_variance_value = max_abs_recursive_variance_from_ok_blocks(text)

    return {
        "strategy": strategy,
        "recursive_timerange": recursive_timerange,
        "recursive_checked": "TRUE",
        "recursive_run_status": run_status,
        "recursive_pair_count": pair_count,
        "recursive_finished_pairs": finished_count,
        "recursive_ok_pairs": ok_count,
        "recursive_failed_pairs": failed_count,
        "recursive_pending_pairs": pending_count,
        "recursive_max_abs_variance": max_abs_variance_value,
        "recursive_created_at": created_at_value(text),
    }


########################################################################################################################################################
# Row combining
########################################################################################################################################################

def apply_lookahead_to_row(row: dict[str, Any], lookahead: dict[str, Any]) -> dict[str, Any]:
    row = dict(row)

    row["Lookahead Timerange"] = lookahead.get("lookahead_timerange", "")
    row["Lookahead Checked"] = lookahead.get("lookahead_checked", "TRUE")
    row["Lookahead Has Bias"] = lookahead.get("lookahead_has_bias", "")
    row["Lookahead Total Signals"] = lookahead.get("lookahead_total_signals", "")

    return {col: row.get(col, "") for col in COLUMNS}


def apply_recursive_to_row(row: dict[str, Any], recursive: dict[str, Any]) -> dict[str, Any]:
    row = dict(row)

    row["Recursive Timerange"] = recursive.get("recursive_timerange", "")
    row["Recursive Checked"] = recursive.get("recursive_checked", "TRUE")
    row["Recursive Run Status"] = recursive.get("recursive_run_status", "")
    row["Recursive Pair Count"] = recursive.get("recursive_pair_count", "")
    row["Recursive Finished Pairs"] = recursive.get("recursive_finished_pairs", "")
    row["Recursive OK Pairs"] = recursive.get("recursive_ok_pairs", "")
    row["Recursive Failed Pairs"] = recursive.get("recursive_failed_pairs", "")
    row["Recursive Pending Pairs"] = recursive.get("recursive_pending_pairs", "")
    row["Recursive Max Abs Variance %"] = recursive.get("recursive_max_abs_variance", "")

    return {col: row.get(col, "") for col in COLUMNS}


def row_strategy_key(row_or_sidecar: dict[str, Any]) -> str:
    return normalize_strategy_key(row_or_sidecar.get("Strategy") or row_or_sidecar.get("strategy") or "")


def sidecar_strategy_key(sidecar: dict[str, Any]) -> str:
    return normalize_strategy_key(sidecar.get("strategy") or "")


def best_sidecar_for_row(row: dict[str, Any], sidecars: List[dict[str, Any]]) -> dict[str, Any] | None:
    if not sidecars:
        return None

    row_key = row_strategy_key(row)

    for sidecar in sidecars:
        if row_key and sidecar_strategy_key(sidecar) == row_key:
            return sidecar

    if len(sidecars) == 1:
        return sidecars[0]

    return None


def combine_rows_with_sidecars(
    hyperopt_rows: List[dict[str, Any]],
    lookahead_rows: List[dict[str, Any]],
    recursive_rows: List[dict[str, Any]],
) -> List[dict[str, Any]]:
    output: List[dict[str, Any]] = []

    for row in hyperopt_rows:
        combined = dict(row)

        lookahead = best_sidecar_for_row(combined, lookahead_rows)
        recursive = best_sidecar_for_row(combined, recursive_rows)

        if lookahead:
            combined = apply_lookahead_to_row(combined, lookahead)

        if recursive:
            combined = apply_recursive_to_row(combined, recursive)

        output.append({col: combined.get(col, "") for col in COLUMNS})

    if not output:
        sidecar_keys = sorted({sidecar_strategy_key(x) for x in lookahead_rows + recursive_rows if sidecar_strategy_key(x)})

        if not sidecar_keys and (lookahead_rows or recursive_rows):
            sidecar_keys = ["__single_sidecar__"]

        for key in sidecar_keys:
            if key == "__single_sidecar__":
                lookahead = lookahead_rows[0] if lookahead_rows else None
                recursive = recursive_rows[0] if recursive_rows else None
            else:
                lookahead = next((x for x in lookahead_rows if sidecar_strategy_key(x) == key), None)
                recursive = next((x for x in recursive_rows if sidecar_strategy_key(x) == key), None)

            strategy = first_non_empty(
                lookahead.get("strategy") if lookahead else "",
                recursive.get("strategy") if recursive else "",
            )

            row = empty_output_row()
            row["Strategy"] = strategy
            row["Created At"] = first_non_empty(
                lookahead.get("lookahead_created_at") if lookahead else "",
                recursive.get("recursive_created_at") if recursive else "",
            )

            if lookahead:
                row = apply_lookahead_to_row(row, lookahead)

            if recursive:
                row = apply_recursive_to_row(row, recursive)

            output.append({col: row.get(col, "") for col in COLUMNS})

    return output


########################################################################################################################################################
# Multi-kind parsing
########################################################################################################################################################

def parse_text_to_groups(text: str) -> Tuple[List[dict[str, Any]], List[dict[str, Any]], List[dict[str, Any]]]:
    clean = clean_text(text)

    if is_hyperopt_text(clean):
        return [parse_hyperopt_extract_text(clean)], [], []

    hyperopt_rows: List[dict[str, Any]] = []
    lookahead_rows: List[dict[str, Any]] = []
    recursive_rows: List[dict[str, Any]] = []

    if is_lookahead_text(clean):
        lookahead_rows.append(parse_lookahead_text(clean))

    if is_recursive_text(clean):
        recursive_rows.append(parse_recursive_text(clean))

    if not lookahead_rows and not recursive_rows:
        hyperopt_rows.append(parse_hyperopt_extract_text(clean))

    return hyperopt_rows, lookahead_rows, recursive_rows


def parse_file_to_groups(path: str | Path) -> Tuple[List[dict[str, Any]], List[dict[str, Any]], List[dict[str, Any]], List[str]]:
    text = read_text_file(path)
    hyperopt_rows, lookahead_rows, recursive_rows = parse_text_to_groups(text)
    kinds: List[str] = []

    if hyperopt_rows:
        kinds.append("hyperopt")
    if lookahead_rows:
        kinds.append("lookahead")
    if recursive_rows:
        kinds.append("recursive")

    return hyperopt_rows, lookahead_rows, recursive_rows, kinds


def parse_any_text(text: str) -> tuple[str, dict[str, Any]]:
    hyperopt_rows, lookahead_rows, recursive_rows = parse_text_to_groups(text)

    if hyperopt_rows:
        return "hyperopt", hyperopt_rows[0]
    if recursive_rows and not lookahead_rows:
        return "recursive", recursive_rows[0]
    if lookahead_rows and not recursive_rows:
        return "lookahead", lookahead_rows[0]
    if lookahead_rows and recursive_rows:
        return "analysis_bundle", {"lookahead": lookahead_rows[0], "recursive": recursive_rows[0]}

    return "hyperopt", parse_hyperopt_extract_text(text)


def parse_any_file(path: str | Path) -> tuple[str, dict[str, Any]]:
    return parse_any_text(read_text_file(path))


########################################################################################################################################################
# Output helpers
########################################################################################################################################################

def values_for_row(row: dict[str, Any]) -> List[Any]:
    return ["" if row.get(col, "") is None else row.get(col, "") for col in COLUMNS]


def tsv_line(row: dict[str, Any]) -> str:
    return "\t".join(str(value) for value in values_for_row(row))


def append_rows_to_csv(rows: List[dict[str, Any]], csv_path: str | Path) -> None:
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    file_exists = csv_path.exists()

    with csv_path.open("a", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=COLUMNS)

        if not file_exists:
            writer.writeheader()

        for row in rows:
            writer.writerow({col: row.get(col, "") for col in COLUMNS})


def import_openpyxl_or_install():
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore

        return Workbook, load_workbook
    except Exception:
        if getattr(sys, "frozen", False):
            raise RuntimeError(
                "openpyxl was not bundled inside the EXE. "
                "Rebuild with: --hidden-import openpyxl --hidden-import et_xmlfile"
            )

        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

    try:
        from openpyxl import Workbook, load_workbook  # type: ignore

        return Workbook, load_workbook
    except Exception as exc:
        raise RuntimeError(
            f"openpyxl is not available in this Python: {sys.executable}. "
            f"Run: {sys.executable} -m pip install openpyxl"
        ) from exc


def append_rows_to_xlsx(rows: List[dict[str, Any]], xlsx_path: str | Path, sheet_name: str = "Raw_Data") -> None:
    Workbook, load_workbook = import_openpyxl_or_install()

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    existing_header = list(first_row[0]) if first_row else []
    is_blank_header = not existing_header or all(value is None for value in existing_header)

    if is_blank_header:
        for idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col)
    else:
        existing_header_trimmed = existing_header[: len(COLUMNS)]

        if existing_header_trimmed != COLUMNS:
            ws.insert_rows(1)

            for idx, col in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=col)

    for row in rows:
        ws.append(values_for_row(row))

    wb.save(xlsx_path)


########################################################################################################################################################
# UI
########################################################################################################################################################

@dataclass
class ParseState:
    rows: List[dict[str, Any]] = field(default_factory=list)
    source_files: List[str] = field(default_factory=list)
    base_rows: List[dict[str, Any]] = field(default_factory=list)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Freqtrade Summary Parser")
        self.root.geometry("1250x1120")
        self.root.minsize(1050, 1000)
        self.root.configure(bg="#111827")

        self.toast_after_id = None
        self.auto_parse_after_id = None
        self.state = ParseState()
        self.multi_file_placeholder = False

        self.hyperopt_file_path = ""
        self.lookahead_file_path = ""
        self.recursive_file_path = ""

        self.setup_style()

        main = tk.Frame(root, bg="#111827")
        main.pack(fill="both", expand=True, padx=14, pady=14)

        title = tk.Label(
            main,
            text="Freqtrade Summary Parser → Excel Row",
            font=("Segoe UI", 18, "bold"),
            fg="#F9FAFB",
            bg="#111827",
        )
        title.pack(anchor="w", pady=(0, 6))

        dnd_status = "Drag & drop: ENABLED" if HAS_DND else "Drag & drop: DISABLED"
        dnd_color = "#10B981" if HAS_DND else "#F59E0B"

        info = tk.Label(
            main,
            text="Only parses raw proof columns. Gates are calculated in Excel formulas.",
            font=("Segoe UI", 10),
            fg="#9CA3AF",
            bg="#111827",
        )
        info.pack(anchor="w", pady=(0, 2))

        dnd_label = tk.Label(
            main,
            text=dnd_status,
            font=("Segoe UI", 10, "bold"),
            fg=dnd_color,
            bg="#111827",
        )
        dnd_label.pack(anchor="w", pady=(0, 12))

        top_bar = tk.Frame(main, bg="#111827")
        top_bar.pack(fill="x", pady=(0, 12))

        left_btns = tk.Frame(top_bar, bg="#111827")
        left_btns.pack(side="left")

        self.make_button(left_btns, "Parse", self.parse_text, "#059669").pack(side="left", padx=(0, 8))
        self.make_button(left_btns, "Copy Headers", self.copy_headers, "#D97706").pack(side="left", padx=(0, 8))
        self.make_button(left_btns, "Copy Row(s)", self.copy_rows, "#7C3AED").pack(side="left", padx=(0, 8))
        self.make_button(left_btns, "Clear", self.clear_all, "#DC2626").pack(side="left", padx=(0, 8))

        right_btns = tk.Frame(top_bar, bg="#111827")
        right_btns.pack(side="right")

        self.make_button(right_btns, "Save CSV", self.save_csv, "#D97706").pack(side="left", padx=(0, 8))
        self.make_button(right_btns, "Append XLSX", self.append_xlsx, "#2563EB").pack(side="left")

        self.file_slots_frame = tk.Frame(main, bg="#111827")
        self.file_slots_frame.pack(fill="x", pady=(0, 12))

        self.hyperopt_slot = self.make_file_slot(
            self.file_slots_frame,
            "1) Hyperopt Extract",
            "Drop/select hyperopt extract/raw output",
            self.select_hyperopt_file,
            "#2563EB",
        )
        self.hyperopt_slot.expected_kind = "hyperopt"
        self.hyperopt_slot.pack(fill="x", pady=(0, 6))

        self.lookahead_slot = self.make_file_slot(
            self.file_slots_frame,
            "2) Lookahead Analysis",
            "Drop/select lookahead-analysis file",
            self.select_lookahead_file,
            "#7C3AED",
        )
        self.lookahead_slot.expected_kind = "lookahead"
        self.lookahead_slot.pack(fill="x", pady=(0, 6))

        self.recursive_slot = self.make_file_slot(
            self.file_slots_frame,
            "3) Recursive Analysis",
            "Drop/select recursive-analysis file",
            self.select_recursive_file,
            "#D97706",
        )
        self.recursive_slot.expected_kind = "recursive"
        self.recursive_slot.pack(fill="x", pady=(0, 6))

        self.ready_status_label = tk.Label(
            main,
            text="Hyperopt: NO | Lookahead: NO | Recursive: NO | Ready: NO",
            font=("Segoe UI", 11, "bold"),
            fg="#F59E0B",
            bg="#111827",
        )
        self.ready_status_label.pack(anchor="w", pady=(0, 10))

        input_label = tk.Label(
            main,
            text="Paste/input box",
            font=("Segoe UI", 11, "bold"),
            fg="#E5E7EB",
            bg="#111827",
        )
        input_label.pack(anchor="w", pady=(0, 6))

        self.input_box = self.make_textbox(main, height=14)
        self.input_box.pack(fill="both", expand=True, pady=(0, 12))

        if HAS_DND:
            for widget in (self.root, self.input_box, self.input_box.text_widget):
                try:
                    widget.drop_target_register(DND_FILES)
                    widget.dnd_bind("<<Drop>>", self.on_file_drop)
                except Exception:
                    pass

        self.input_box.text_widget.bind("<<Paste>>", self.on_input_changed_event, add="+")
        self.input_box.text_widget.bind("<KeyRelease>", self.on_input_changed_event, add="+")
        self.input_box.text_widget.bind("<ButtonRelease-3>", self.on_input_changed_event, add="+")
        self.input_box.text_widget.bind("<Control-v>", self.on_input_changed_event, add="+")
        self.input_box.text_widget.bind("<Control-V>", self.on_input_changed_event, add="+")

        headers_label = tk.Label(
            main,
            text="Headers — click inside box to copy",
            font=("Segoe UI", 11, "bold"),
            fg="#E5E7EB",
            bg="#111827",
        )
        headers_label.pack(anchor="w", pady=(0, 6))

        self.headers_box = self.make_textbox(main, height=4, readonly=True)
        self.headers_box.pack(fill="x", expand=False, pady=(0, 12))
        self.bind_copy_box(self.headers_box, self.copy_headers)

        row_label = tk.Label(
            main,
            text="Excel-ready final parsed row(s) — click inside box to copy",
            font=("Segoe UI", 11, "bold"),
            fg="#E5E7EB",
            bg="#111827",
        )
        row_label.pack(anchor="w", pady=(0, 6))

        self.row_box = self.make_textbox(main, height=8, readonly=True)
        self.row_box.pack(fill="both", expand=True, pady=(0, 12))
        self.bind_copy_box(self.row_box, self.copy_rows)

        self.toast = tk.Label(
            main,
            text="",
            font=("Segoe UI", 10, "bold"),
            fg="#F9FAFB",
            bg="#1F2937",
            bd=0,
            padx=14,
            pady=8,
        )
        self.toast.place_forget()

        self.display_rows([])
        self.update_file_slot_labels()

    def setup_style(self):
        try:
            style = ttk.Style()
            try:
                style.theme_use("clam")
            except Exception:
                pass
        except Exception:
            pass

    def make_button(self, parent, text, command, bg):
        return tk.Button(
            parent,
            text=text,
            command=command,
            font=("Segoe UI", 10, "bold"),
            fg="#FFFFFF",
            bg=bg,
            activebackground=bg,
            activeforeground="#FFFFFF",
            relief="flat",
            bd=0,
            padx=14,
            pady=8,
            cursor="hand2",
        )

    def make_file_slot(self, parent, title, placeholder, command, button_color):
        outer = tk.Frame(parent, bg="#1F2937", highlightthickness=1, highlightbackground="#374151")

        title_label = tk.Label(
            outer,
            text=title,
            font=("Segoe UI", 10, "bold"),
            fg="#E5E7EB",
            bg="#1F2937",
            width=22,
            anchor="w",
        )
        title_label.pack(side="left", padx=(10, 8), pady=8)

        status_label = tk.Label(
            outer,
            text="EMPTY",
            font=("Segoe UI", 10, "bold"),
            fg="#F59E0B",
            bg="#1F2937",
            width=8,
            anchor="w",
        )
        status_label.pack(side="left", padx=(0, 8), pady=8)

        path_label = tk.Label(
            outer,
            text=placeholder,
            font=("Consolas", 9),
            fg="#9CA3AF",
            bg="#1F2937",
            anchor="w",
        )
        path_label.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=8)

        button = self.make_button(outer, "Select", command, button_color)
        button.pack(side="right", padx=(0, 8), pady=6)

        clear_button = self.make_button(outer, "Clear", lambda: self.clear_file_slot(outer), "#DC2626")
        clear_button.pack(side="right", padx=(0, 8), pady=6)

        outer.title_label = title_label
        outer.status_label = status_label
        outer.path_label = path_label
        outer.placeholder = placeholder
        outer.expected_kind = ""

        if HAS_DND:
            for widget in (outer, title_label, status_label, path_label):
                try:
                    widget.drop_target_register(DND_FILES)
                    widget.dnd_bind("<<Drop>>", lambda event, slot=outer: self.on_slot_drop(event, slot))
                except Exception:
                    pass

        return outer

    def make_textbox(self, parent, height=8, readonly=False):
        outer = tk.Frame(parent, bg="#1F2937", highlightthickness=1, highlightbackground="#374151")

        text = tk.Text(
            outer,
            wrap="none",
            height=height,
            font=("Consolas", 10),
            bg="#0F172A",
            fg="#E5E7EB",
            insertbackground="#F9FAFB",
            relief="flat",
            bd=0,
            undo=not readonly,
            padx=10,
            pady=10,
        )

        xscroll = tk.Scrollbar(outer, orient="horizontal", command=text.xview)
        yscroll = tk.Scrollbar(outer, orient="vertical", command=text.yview)

        text.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)

        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        text.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        if readonly:
            text.configure(state="disabled")

        outer.text_widget = text
        return outer

    def bind_copy_box(self, outer, callback):
        outer.bind("<Button-1>", lambda event: self.handle_copy_click(callback))
        outer.text_widget.bind("<Button-1>", lambda event: self.handle_copy_click(callback))
        outer.text_widget.bind("<ButtonRelease-1>", lambda event: "break")
        outer.text_widget.bind("<B1-Motion>", lambda event: "break")
        outer.text_widget.bind("<Double-Button-1>", lambda event: "break")
        outer.text_widget.bind("<Triple-Button-1>", lambda event: "break")

    def handle_copy_click(self, callback):
        callback()
        return "break"

    def get_textbox_text(self, outer):
        return outer.text_widget.get("1.0", "end").strip()

    def set_textbox_text(self, outer, value):
        widget = outer.text_widget
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", value)

        if outer in (self.headers_box, self.row_box):
            widget.configure(state="disabled")

    def clear_textbox(self, outer):
        widget = outer.text_widget
        widget.configure(state="normal")
        widget.delete("1.0", "end")

        if outer in (self.headers_box, self.row_box):
            widget.configure(state="disabled")

    def show_toast(self, message, bg="#1F2937"):
        self.toast.configure(text=message, bg=bg)
        self.toast.place(relx=1.0, rely=1.0, x=-20, y=-20, anchor="se")

        if self.toast_after_id is not None:
            self.root.after_cancel(self.toast_after_id)

        self.toast_after_id = self.root.after(5000, self.hide_toast)

    def hide_toast(self):
        self.toast.place_forget()
        self.toast_after_id = None

    def copy_to_clipboard(self, text, label="Copied"):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        self.show_toast(label, bg="#065F46")

    def update_ready_status(self, rows: List[dict[str, Any]]):
        def row_has_hyperopt(row: dict[str, Any]) -> bool:
            return any(
                str(row.get(col, "")).strip()
                for col in [
                    "Hyperopt Seed",
                    "Epochs Count",
                    "Strategy Version",
                    "Hyperopt Loss",
                    "Timerange Group",
                    "Train Timerange",
                    "Valid Timerange",
                    "Test Timerange",
                ]
            )

        def row_has_lookahead(row: dict[str, Any]) -> bool:
            return str(row.get("Lookahead Checked", "")).strip().upper() == "TRUE"

        def row_has_recursive(row: dict[str, Any]) -> bool:
            return str(row.get("Recursive Checked", "")).strip().upper() == "TRUE"

        has_hyperopt = any(row_has_hyperopt(row) for row in rows)
        has_lookahead = any(row_has_lookahead(row) for row in rows)
        has_recursive = any(row_has_recursive(row) for row in rows)

        ready = any(
            row_has_hyperopt(row) and row_has_lookahead(row) and row_has_recursive(row)
            for row in rows
        )

        self.ready_status_label.configure(
            text=(
                f"Hyperopt: {'YES' if has_hyperopt else 'NO'} | "
                f"Lookahead: {'YES' if has_lookahead else 'NO'} | "
                f"Recursive: {'YES' if has_recursive else 'NO'} | "
                f"Ready to copy final parsed row: {'YES' if ready else 'NO'}"
            ),
            fg="#10B981" if ready else "#F59E0B",
        )

    def update_file_slot_labels(self):
        slots = [
            (self.hyperopt_slot, self.hyperopt_file_path),
            (self.lookahead_slot, self.lookahead_file_path),
            (self.recursive_slot, self.recursive_file_path),
        ]

        for slot, path in slots:
            if path:
                slot.status_label.configure(text="READY", fg="#10B981")
                slot.path_label.configure(text=path, fg="#E5E7EB")
            else:
                slot.status_label.configure(text="EMPTY", fg="#F59E0B")
                slot.path_label.configure(text=slot.placeholder, fg="#9CA3AF")

        self.update_ready_status(self.state.rows)

    def apply_manual_overrides_to_rows(self, rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
        return [{col: row.get(col, "") for col in COLUMNS} for row in rows]

    def detected_kinds_for_file(self, path: str) -> List[str]:
        return detected_kinds_for_text(read_text_file(path))

    def set_slot_path(self, kind: str, path: str, validate: bool = True):
        detected_kinds: List[str] = []

        if validate:
            try:
                detected_kinds = self.detected_kinds_for_file(path)
            except Exception as exc:
                self.show_toast(f"Could not read file: {exc}", bg="#991B1B")
                return

            if kind not in detected_kinds:
                self.show_toast(
                    f"Wrong file type. This slot expects {kind}, but file is {'+'.join(detected_kinds)}.",
                    bg="#991B1B",
                )
                return

        if kind == "hyperopt":
            self.hyperopt_file_path = path
        elif kind == "lookahead":
            self.lookahead_file_path = path
            if "recursive" in detected_kinds and not self.recursive_file_path:
                self.recursive_file_path = path
        elif kind == "recursive":
            self.recursive_file_path = path
            if "lookahead" in detected_kinds and not self.lookahead_file_path:
                self.lookahead_file_path = path

        self.update_file_slot_labels()
        self.parse_slot_files()

    def clear_file_slot(self, slot):
        if slot is self.hyperopt_slot:
            self.hyperopt_file_path = ""
        elif slot is self.lookahead_slot:
            self.lookahead_file_path = ""
        elif slot is self.recursive_slot:
            self.recursive_file_path = ""

        self.update_file_slot_labels()
        self.parse_slot_files()

    def select_file_for_slot(self, kind: str):
        file_path = filedialog.askopenfilename(
            title=f"Select {kind} file",
            filetypes=[
                ("Freqtrade text files", "*.txt *.log *.ds *.md"),
                ("All files", "*.*"),
            ],
        )

        if file_path:
            self.set_slot_path(kind, file_path, validate=True)

    def select_hyperopt_file(self):
        self.select_file_for_slot("hyperopt")

    def select_lookahead_file(self):
        self.select_file_for_slot("lookahead")

    def select_recursive_file(self):
        self.select_file_for_slot("recursive")

    def on_slot_drop(self, event, slot):
        try:
            paths = self._extract_dropped_filepaths(event.data)

            if not paths:
                return "break"

            path = paths[0]
            expected_kind = getattr(slot, "expected_kind", "")

            if expected_kind not in {"hyperopt", "lookahead", "recursive"}:
                self.show_toast("Unknown slot type.", bg="#991B1B")
                return "break"

            self.set_slot_path(expected_kind, path, validate=True)

        except Exception as exc:
            self.show_toast(f"Slot drop error: {exc}", bg="#991B1B")

        return "break"

    def parse_slot_files(self):
        paths = unique_paths([p for p in [self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path] if p])

        if not paths:
            self.state = ParseState()
            self.display_rows([])
            return

        self.parse_files(paths, from_slots=True)

    def schedule_auto_parse(self):
        if self.auto_parse_after_id is not None:
            self.root.after_cancel(self.auto_parse_after_id)

        self.auto_parse_after_id = self.root.after(350, self.auto_parse_now)

    def auto_parse_now(self):
        self.auto_parse_after_id = None

        if any([self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path]):
            return

        text = self.get_textbox_text(self.input_box)

        if not text:
            return

        try:
            hyperopt_rows, lookahead_rows, recursive_rows = parse_text_to_groups(text)
            rows = combine_rows_with_sidecars(hyperopt_rows, lookahead_rows, recursive_rows)
            rows = self.apply_manual_overrides_to_rows(rows)
            self.state = ParseState(rows=rows, source_files=[], base_rows=rows)
            self.display_rows(rows)

        except Exception:
            pass

    def on_input_changed_event(self, event=None):
        self.multi_file_placeholder = False
        self.schedule_auto_parse()

    def _extract_dropped_filepaths(self, data: str) -> List[str]:
        try:
            paths = list(self.root.tk.splitlist(data))
        except Exception:
            paths = [data]

        cleaned = []

        for path in paths:
            path = path.strip().strip("{}").strip()

            if path:
                cleaned.append(path)

        return cleaned

    def on_file_drop(self, event):
        try:
            paths = self._extract_dropped_filepaths(event.data)

            if paths:
                self.parse_files(paths, from_slots=False)

        except Exception as exc:
            self.show_toast(f"Drop load error: {exc}", bg="#991B1B")

        return "break"

    def parse_files(self, paths: Iterable[str], from_slots: bool = False):
        clean_paths = unique_paths([str(Path(path)) for path in paths if str(path).strip()])

        hyperopt_rows: List[dict[str, Any]] = []
        lookahead_rows: List[dict[str, Any]] = []
        recursive_rows: List[dict[str, Any]] = []
        errors: List[str] = []

        detected_hyperopt = ""
        detected_lookahead = ""
        detected_recursive = ""

        for file_path in clean_paths:
            try:
                file_hyperopt_rows, file_lookahead_rows, file_recursive_rows, kinds = parse_file_to_groups(file_path)

                hyperopt_rows.extend(file_hyperopt_rows)
                lookahead_rows.extend(file_lookahead_rows)
                recursive_rows.extend(file_recursive_rows)

                if "hyperopt" in kinds:
                    detected_hyperopt = file_path
                if "lookahead" in kinds:
                    detected_lookahead = file_path
                if "recursive" in kinds:
                    detected_recursive = file_path

            except Exception as exc:
                errors.append(f"{Path(file_path).name}: {exc}")

        if not from_slots:
            if detected_hyperopt:
                self.hyperopt_file_path = detected_hyperopt
            if detected_lookahead:
                self.lookahead_file_path = detected_lookahead
            if detected_recursive:
                self.recursive_file_path = detected_recursive

            slot_paths = unique_paths([p for p in [self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path] if p])

            if slot_paths and sorted(slot_paths) != sorted(clean_paths):
                self.update_file_slot_labels()
                self.parse_files(slot_paths, from_slots=True)
                return

        rows = combine_rows_with_sidecars(hyperopt_rows, lookahead_rows, recursive_rows)
        rows = self.apply_manual_overrides_to_rows(rows)

        if not rows:
            message = "No valid rows parsed"

            if errors:
                message += f": {errors[0]}"

            self.show_toast(message, bg="#991B1B")
            self.update_file_slot_labels()
            return

        self.state = ParseState(rows=rows, source_files=clean_paths, base_rows=rows)

        if len(clean_paths) == 1:
            self.multi_file_placeholder = False

            try:
                self.set_textbox_text(self.input_box, read_text_file(clean_paths[0]))
            except Exception:
                self.set_textbox_text(self.input_box, f"Loaded file: {clean_paths[0]}")
        else:
            self.multi_file_placeholder = True
            file_list = "\n".join(clean_paths)

            self.set_textbox_text(
                self.input_box,
                (
                    "Loaded multiple files. Parsed rows are shown below.\n\n"
                    f"Hyperopt files: {len(hyperopt_rows)}\n"
                    f"Lookahead files: {len(lookahead_rows)}\n"
                    f"Recursive files: {len(recursive_rows)}\n\n"
                    f"{file_list}"
                ),
            )

        self.update_file_slot_labels()
        self.display_rows(rows)

        if errors:
            self.show_toast(f"Parsed {len(rows)} row(s), {len(errors)} failed. First: {errors[0]}", bg="#92400E")
        else:
            self.show_toast(f"Parsed {len(rows)} row(s)", bg="#065F46")

    def parse_text(self):
        if any([self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path]):
            self.parse_slot_files()
            return

        text = self.get_textbox_text(self.input_box)

        if not text:
            self.show_toast("Drop files into the 3 slots or paste text into the input box", bg="#92400E")
            return

        try:
            hyperopt_rows, lookahead_rows, recursive_rows = parse_text_to_groups(text)
            rows = combine_rows_with_sidecars(hyperopt_rows, lookahead_rows, recursive_rows)
            rows = self.apply_manual_overrides_to_rows(rows)

        except Exception as exc:
            self.show_toast(f"Parse error: {exc}", bg="#991B1B")
            return

        self.state = ParseState(rows=rows, source_files=[], base_rows=rows)
        self.display_rows(rows)

        strategy = rows[0].get("Strategy", "") or "unknown strategy"
        self.show_toast(f"Parsed successfully: {strategy}", bg="#065F46")

    def display_rows(self, rows: List[dict[str, Any]]):
        headers = "\t".join(COLUMNS)
        row_text = "\n".join(tsv_line(row) for row in rows)

        self.set_textbox_text(self.headers_box, headers)
        self.set_textbox_text(self.row_box, row_text)
        self.update_ready_status(rows)

    def copy_headers(self):
        headers = self.get_textbox_text(self.headers_box) or "\t".join(COLUMNS)

        if not headers:
            self.show_toast("Nothing to copy in headers box", bg="#92400E")
            return

        self.copy_to_clipboard(headers, "Headers copied")

    def copy_rows(self):
        if any([self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path]):
            self.parse_slot_files()
        elif not self.state.rows:
            self.parse_text()

        row = self.get_textbox_text(self.row_box)

        if not row:
            self.show_toast("Nothing to copy in row box", bg="#92400E")
            return

        self.copy_to_clipboard(row, "Excel row(s) copied")

    def save_csv(self):
        if any([self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path]):
            self.parse_slot_files()
        elif not self.state.rows:
            self.parse_text()

        if not self.state.rows:
            self.show_toast("No parsed rows to save", bg="#92400E")
            return

        path = filedialog.asksaveasfilename(
            title="Save / append CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )

        if not path:
            return

        try:
            append_rows_to_csv(self.state.rows, path)
        except Exception as exc:
            self.show_toast(f"CSV save error: {exc}", bg="#991B1B")
            return

        self.show_toast(f"Saved/appended {len(self.state.rows)} row(s) to CSV", bg="#065F46")

    def append_xlsx(self):
        if any([self.hyperopt_file_path, self.lookahead_file_path, self.recursive_file_path]):
            self.parse_slot_files()
        elif not self.state.rows:
            self.parse_text()

        if not self.state.rows:
            self.show_toast("No parsed rows to append", bg="#92400E")
            return

        path = filedialog.asksaveasfilename(
            title="Save / append XLSX Raw_Data",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )

        if not path:
            return

        try:
            append_rows_to_xlsx(self.state.rows, path)
        except Exception as exc:
            self.show_toast(f"XLSX append error: {exc}", bg="#991B1B")
            return

        self.show_toast(f"Saved/appended {len(self.state.rows)} row(s) to XLSX Raw_Data", bg="#065F46")

    def clear_all(self):
        if self.auto_parse_after_id is not None:
            self.root.after_cancel(self.auto_parse_after_id)
            self.auto_parse_after_id = None

        self.multi_file_placeholder = False
        self.state = ParseState()
        self.hyperopt_file_path = ""
        self.lookahead_file_path = ""
        self.recursive_file_path = ""

        self.clear_textbox(self.input_box)
        self.display_rows([])
        self.update_file_slot_labels()
        self.show_toast("Cleared", bg="#991B1B")


########################################################################################################################################################
# CLI fallback
########################################################################################################################################################

def cli_main(argv: List[str]) -> int:
    parser = argparse.ArgumentParser(description="Parse Freqtrade hyperopt/lookahead/recursive metadata to Excel-ready rows.")
    parser.add_argument("files", nargs="*", help="Hyperopt extract plus optional Lookahead/Recursive analysis files. Omit files to launch UI.")
    parser.add_argument("--csv", default="", help="Optional CSV path to append parsed rows.")
    parser.add_argument("--xlsx", default="", help="Optional XLSX path to append parsed rows to Raw_Data sheet.")
    parser.add_argument("--json", action="store_true", help="Print JSON instead of TSV.")

    args = parser.parse_args(argv)

    if not args.files:
        return launch_ui()

    hyperopt_rows: List[dict[str, Any]] = []
    lookahead_rows: List[dict[str, Any]] = []
    recursive_rows: List[dict[str, Any]] = []

    for path in args.files:
        file_hyperopt_rows, file_lookahead_rows, file_recursive_rows, _kinds = parse_file_to_groups(path)
        hyperopt_rows.extend(file_hyperopt_rows)
        lookahead_rows.extend(file_lookahead_rows)
        recursive_rows.extend(file_recursive_rows)

    rows = combine_rows_with_sidecars(hyperopt_rows, lookahead_rows, recursive_rows)

    if args.json:
        print(json.dumps(rows, indent=2, default=str))
    else:
        print("\t".join(COLUMNS))

        for row in rows:
            print(tsv_line(row))

    if args.csv:
        append_rows_to_csv(rows, args.csv)

    if args.xlsx:
        append_rows_to_xlsx(rows, args.xlsx)

    return 0


def launch_ui() -> int:
    root = TkinterDnD.Tk() if HAS_DND and TkinterDnD is not None else tk.Tk()
    App(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(cli_main(sys.argv[1:]))
